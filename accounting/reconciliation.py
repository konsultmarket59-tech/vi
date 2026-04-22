"""
Сверка взаиморасчётов по клиентам.

Создаёт и обновляет XLSX-файл «Сверка_{Клиент}_{year}.xlsx» на Яндекс.Диске.
Каждый клиент — отдельный лист.
Столбцы: Месяц | ДС | ТЗ | Акт № | Дата акта | Сумма акта | Оплачено | Дата оплаты | Баланс

Скрипт только добавляет строки. Оплату пользователь вносит вручную.
"""

import io
import os
import tempfile
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
import requests

YADISK_API = 'https://cloud-api.yandex.net/v1/disk'

HEADERS = [
    'Месяц', 'ДС', 'ТЗ', 'Акт №', 'Дата акта',
    'Сумма акта, ₽', 'Оплачено, ₽', 'Дата оплаты', 'Баланс, ₽',
]

COL_WIDTHS = [12, 6, 6, 8, 12, 16, 14, 14, 14]


# ---------------------------------------------------------------------------
# Яндекс.Диск helpers
# ---------------------------------------------------------------------------

def _ya_headers(token):
    return {'Authorization': f'OAuth {token}'}


def ensure_yadisk_folder(token, path):
    parts = [p for p in path.strip('/').split('/') if p]
    cur = ''
    for part in parts:
        cur = f'{cur}/{part}' if cur else part
        r = requests.put(
            f'{YADISK_API}/resources',
            params={'path': cur},
            headers=_ya_headers(token),
            timeout=30,
        )
        if r.status_code not in (201, 409):
            raise RuntimeError(f'mkdir {cur}: {r.status_code} {r.text[:200]}')


def upload_file_to_yadisk(token, local_path, remote_path):
    r = requests.get(
        f'{YADISK_API}/resources/upload',
        params={'path': remote_path, 'overwrite': 'true'},
        headers=_ya_headers(token),
        timeout=30,
    )
    r.raise_for_status()
    href = r.json()['href']
    with open(local_path, 'rb') as fh:
        put = requests.put(href, data=fh, timeout=300)
    if put.status_code >= 400:
        raise RuntimeError(f'PUT {put.status_code}: {put.text[:200]}')
    print(f'  ✓ сверка загружена: {remote_path}')


def download_file_from_yadisk(token, remote_path):
    r = requests.get(
        f'{YADISK_API}/resources/download',
        params={'path': remote_path},
        headers=_ya_headers(token),
        timeout=30,
    )
    if r.status_code == 404:
        return None
    r.raise_for_status()
    data = requests.get(r.json()['href'], timeout=60)
    data.raise_for_status()
    return data.content


# ---------------------------------------------------------------------------
# Стили
# ---------------------------------------------------------------------------

def _header_style():
    fill = PatternFill('solid', fgColor='1C4587')
    font = Font(color='FFFFFF', bold=True, size=10)
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return fill, font, align, border


def _data_border():
    thin = Side(style='thin', color='CCCCCC')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _balance_formula(row_num):
    return f'=F{row_num}-G{row_num}'


# ---------------------------------------------------------------------------
# Создание / обновление листа клиента
# ---------------------------------------------------------------------------

def _ensure_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    _write_header_row(ws)
    return ws


def _write_header_row(ws):
    fill, font, align, border = _header_style()
    for col, (title, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = 'A2'


def _row_already_exists(ws, period_str, ds, tz, act_num):
    """Проверяем, не добавлена ли уже эта строка (чтобы не дублировать)."""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if (str(row[0]) == period_str and
                str(row[1]) == str(ds) and
                str(row[3]) == str(act_num)):
            return True
    return False


def add_act_row(ws, period_str, ds, tz, act_num, act_date, amount):
    if _row_already_exists(ws, period_str, ds, tz, act_num):
        print(f'    Строка уже есть: {period_str} ДС{ds} Акт№{act_num} — пропускаем')
        return

    next_row = ws.max_row + 1
    border = _data_border()
    center = Alignment(horizontal='center', vertical='center')
    money_fmt = '#,##0.00'

    values = [
        period_str,      # A Месяц
        str(ds),         # B ДС
        str(tz) if tz else '',  # C ТЗ
        str(act_num),    # D Акт №
        act_date,        # E Дата акта
        amount,          # F Сумма акта
        None,            # G Оплачено (вручную)
        None,            # H Дата оплаты (вручную)
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.border = border
        cell.alignment = center
        if col in (6, 7):
            cell.number_format = money_fmt

    # Баланс = Сумма - Оплачено (формула)
    bal_cell = ws.cell(row=next_row, column=9,
                       value=f'=F{next_row}-IF(G{next_row}="",0,G{next_row})')
    bal_cell.border = border
    bal_cell.alignment = center
    bal_cell.number_format = money_fmt
    bal_cell.font = Font(bold=True, color='CC0000')


# ---------------------------------------------------------------------------
# Главная функция
# ---------------------------------------------------------------------------

def update_reconciliation(token, accounting_folder, generated_acts, period_year):
    """
    generated_acts: список dict из generate_monthly_acts.run()
      {client, type, ds, tz (optional), total, act_num, path}
    """
    if not generated_acts:
        print('Нет актов для обновления сверки.')
        return

    # Группируем по клиенту
    by_client = {}
    for act in generated_acts:
        by_client.setdefault(act['client'], []).append(act)

    for client_name, acts in by_client.items():
        remote_path = (
            f'{accounting_folder}/{client_name}/'
            f'Сверка_{client_name}_{period_year}.xlsx'
        )

        # Скачать существующий файл или создать новый
        existing_bytes = download_file_from_yadisk(token, remote_path)
        if existing_bytes:
            wb = openpyxl.load_workbook(io.BytesIO(existing_bytes))
        else:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        ws = _ensure_sheet(wb, client_name[:31])  # лист <= 31 символ
        act_date_str = date(period_year,
                            acts[0].get('month', date.today().month),
                            1).strftime('%d.%m.%Y')
        # Если month не передан, берём из path (YYYY-MM)
        period_str = acts[0]['path'].split('/')[-2] if '/' in acts[0]['path'] else ''

        for act in acts:
            add_act_row(
                ws,
                period_str=period_str or act.get('period', ''),
                ds=act['ds'],
                tz=act.get('tz', ''),
                act_num=act['act_num'],
                act_date=date.today().strftime('%d.%m.%Y'),
                amount=act['total'],
            )

        # Итоговая строка с суммами (динамическая)
        last_row = ws.max_row
        summary_row = last_row + 2
        ws.cell(row=summary_row, column=5, value='ИТОГО:').font = Font(bold=True)
        ws.cell(row=summary_row, column=6,
                value=f'=SUM(F2:F{last_row})').number_format = '#,##0.00'
        ws.cell(row=summary_row, column=6).font = Font(bold=True)
        ws.cell(row=summary_row, column=7,
                value=f'=SUM(G2:G{last_row})').number_format = '#,##0.00'
        ws.cell(row=summary_row, column=9,
                value=f'=SUM(I2:I{last_row})').number_format = '#,##0.00'
        ws.cell(row=summary_row, column=9).font = Font(bold=True, color='CC0000')

        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)

        ensure_yadisk_folder(
            token,
            f'{accounting_folder}/{client_name}'
        )
        upload_file_to_yadisk(token, tmp.name, remote_path)
        Path(tmp.name).unlink(missing_ok=True)
