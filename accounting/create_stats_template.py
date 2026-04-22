"""
Создаёт пустой Excel-шаблон для ввода данных статистики.
Запускается один раз для получения файла-образца.

Использование:
  python accounting/create_stats_template.py
  → создаёт Шаблон_статистика_YYYY-MM.xlsx в текущей папке
"""

import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.worksheet.datavalidation import DataValidation


def _hdr(ws, row, col, text, width=None):
    fill = PatternFill('solid', fgColor='1C4587')
    font = Font(color='FFFFFF', bold=True, size=10)
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = fill
    cell.font = font
    cell.alignment = align
    cell.border = border
    if width:
        ws.column_dimensions[cell.column_letter].width = width
    return cell


def _example(ws, row, col, text, comment=None):
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell = ws.cell(row=row, column=col, value=text)
    cell.border = border
    cell.font = Font(color='555555', italic=True, size=10)
    cell.alignment = Alignment(vertical='center')
    return cell


def build_template(output_path):
    wb = openpyxl.Workbook()

    # ── Лист 1: СММ посты ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'СММ посты'

    ws1.merge_cells('A1:G1')
    ws1['A1'] = ('Лист «СММ посты» — ежемесячные посты для актов по Доп. соглашению №1. '
                 'Стоимость можно оставить пустой — подставится автоматически по тарифу.')
    ws1['A1'].font = Font(italic=True, size=9, color='555555')
    ws1['A1'].alignment = Alignment(wrap_text=True)
    ws1.row_dimensions[1].height = 28

    headers1 = [
        ('Клиент', 22),
        ('Договор №', 12),
        ('ДС №', 8),
        ('Дата публикации', 18),
        ('Тип контента', 32),
        ('Ссылка на пост', 40),
        ('Стоимость, ₽ (авто если пусто)', 22),
    ]
    for col, (title, width) in enumerate(headers1, 1):
        _hdr(ws1, 2, col, title, width)
    ws1.row_dimensions[2].height = 36
    ws1.freeze_panes = 'A3'

    # Валидация: Тип контента
    content_types = ('"Фото + текст,Инфографика + текст,Инфографика + фото + текст,'
                     'Статья,Генеративное видео + текст,Реальное видео + сториз + текст,'
                     'Реальное видео + текст,Другое"')
    dv1 = DataValidation(type='list', formula1=content_types, showDropDown=False)
    ws1.add_data_validation(dv1)
    dv1.add('E3:E1000')

    # Примеры строк
    examples1 = [
        ['Болдино LIFE (Павлов)', '14', '1', '01.03.2026',
         'Фото + текст', 'https://vk.com/wall-231359205_XX', ''],
        ['Болдино LIFE (Павлов)', '14', '1', '05.03.2026',
         'Реальное видео + сториз + текст', 'https://vk.com/wall-231359205_YY', ''],
    ]
    for r_idx, row_data in enumerate(examples1, 3):
        for c_idx, val in enumerate(row_data, 1):
            _example(ws1, r_idx, c_idx, val)

    # ── Лист 2: Услуги ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Услуги')

    ws2.merge_cells('A1:F1')
    ws2['A1'] = ('Лист «Услуги» — фиксированные услуги для актов по Доп. соглашению №2 '
                 '(Яндекс.Директ, сайт, консультации и т.д.).')
    ws2['A1'].font = Font(italic=True, size=9, color='555555')
    ws2['A1'].alignment = Alignment(wrap_text=True)
    ws2.row_dimensions[1].height = 28

    headers2 = [
        ('Клиент', 26),
        ('Договор №', 12),
        ('ДС №', 8),
        ('ТЗ №', 8),
        ('Наименование услуги', 42),
        ('Стоимость, ₽', 16),
    ]
    for col, (title, width) in enumerate(headers2, 1):
        _hdr(ws2, 2, col, title, width)
    ws2.row_dimensions[2].height = 36
    ws2.freeze_panes = 'A3'

    examples2 = [
        ['Болдино LIFE (Павлов)', '14', '2', '7',
         'Яндекс.Директ: создание и ведение рекламных кампаний', 20000],
        ['Болдино LIFE (Павлов)', '14', '2', '7',
         'Доработки сайта boldino59.ru', 25000],
        ['Болдино LIFE (Павлов)', '14', '2', '7',
         'Информационно-консультационные услуги', 5000],
    ]
    for r_idx, row_data in enumerate(examples2, 3):
        for c_idx, val in enumerate(row_data, 1):
            _example(ws2, r_idx, c_idx, val)

    # ── Лист 3: ОРД статистика ────────────────────────────────────────────────
    ws3 = wb.create_sheet('ОРД статистика')

    ws3.merge_cells('A1:K1')
    ws3['A1'] = ('Лист «ОРД статистика» — данные для шаблона ВК ОРД. '
                 'ERID берётся из раздела «Креативы» в ord.vk.com.')
    ws3['A1'].font = Font(italic=True, size=9, color='555555')
    ws3['A1'].alignment = Alignment(wrap_text=True)
    ws3.row_dimensions[1].height = 28

    headers3 = [
        ('ERID (токен)', 20),
        ('Название креатива', 30),
        ('Площадка', 28),
        ('Показов', 12),
        ('Оплачено показов', 16),
        ('Период с', 14),
        ('Период по', 14),
        ('Тип события', 18),
        ('Стоимость события, ₽', 20),
        ('Сумма с НДС, ₽', 16),
        ('Ставка НДС, %', 14),
    ]
    for col, (title, width) in enumerate(headers3, 1):
        _hdr(ws3, 2, col, title, width)
    ws3.row_dimensions[2].height = 40
    ws3.freeze_panes = 'A3'

    examples3 = [
        ['2VtzqwsYMFk', 'Услуги девелопера',
         'Болдино Лайф | Коттеджный поселок Пермь',
         201, 201, '18.03.2026', '31.03.2026', 'Фиксированная', '', 5000, ''],
    ]
    for r_idx, row_data in enumerate(examples3, 3):
        for c_idx, val in enumerate(row_data, 1):
            _example(ws3, r_idx, c_idx, val)

    # ── Лист 4: Справка ────────────────────────────────────────────────────────
    ws4 = wb.create_sheet('Тарифы и справка')
    ws4['A1'] = 'Тарифная сетка СММ (актуальна на дату создания шаблона)'
    ws4['A1'].font = Font(bold=True, size=12)

    rates = [
        ('Фото + текст', 1400),
        ('Инфографика + текст', 1400),
        ('Инфографика + фото + текст', 1400),
        ('Статья', 2500),
        ('Генеративное видео + текст', 5000),
        ('Реальное видео + сториз + текст', 8000),
        ('Реальное видео + текст', 8000),
    ]
    _hdr(ws4, 2, 1, 'Тип контента', 36)
    _hdr(ws4, 2, 2, 'Стоимость, ₽', 16)
    for r_idx, (name, cost) in enumerate(rates, 3):
        ws4.cell(row=r_idx, column=1, value=name).font = Font(size=10)
        ws4.cell(row=r_idx, column=2, value=cost).font = Font(size=10)

    ws4.cell(row=12, column=1,
             value='* Нестандартные позиции (коллаборации и т.д.) '
                   'вносятся вручную в колонку «Стоимость, ₽» листа «СММ посты».'
             ).font = Font(italic=True, size=9, color='555555')

    wb.save(output_path)
    print(f'Шаблон создан: {output_path}')


if __name__ == '__main__':
    today = date.today()
    # Шаблон для следующего месяца
    if today.month == 12:
        next_year, next_month = today.year + 1, 1
    else:
        next_year, next_month = today.year, today.month + 1

    filename = f'Шаблон_статистика_{next_year}-{next_month:02d}.xlsx'
    output = Path(filename)
    build_template(str(output))
