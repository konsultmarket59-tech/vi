import os
import io
import datetime
import requests
import openpyxl
from pathlib import Path
from anthropic import Anthropic

YADISK_API = "https://cloud-api.yandex.net/v1/disk"

MONTHS_RU = [
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]

# Column name aliases — add synonyms as needed
_COL_ALIASES = {
    "topic":       ["тема", "topic", "заголовок", "название"],
    "description": ["описание", "description", "бриф", "brief"],
    "contacts":    ["контакты", "contacts", "контактные данные"],
    "key_aspects": ["ключевые аспекты", "аспекты", "key aspects", "фокус"],
    "folder":      ["папка", "folder", "месяц", "month"],
    "status":      ["статус", "status", "готово", "done"],
    "result":      ["результат", "result", "ссылка", "link", "url"],
}


def _match_col(header: str, field: str) -> bool:
    return any(alias in header.lower() for alias in _COL_ALIASES[field])


def load_context() -> str:
    context_dir = Path(__file__).parent / "context"
    parts = []
    for path in sorted(context_dir.glob("*.md")):
        if path.name == "README.md":
            continue
        text = path.read_text(encoding="utf-8").strip()
        if text:
            parts.append(f"### {path.stem}\n\n{text}")
    return "\n\n---\n\n".join(parts)


def load_skill() -> str:
    skill_path = Path(__file__).parent / "content_funnel_skill.md"
    return skill_path.read_text(encoding="utf-8")


class YaDiskClient:
    def __init__(self, token: str):
        self.headers = {"Authorization": f"OAuth {token}"}

    def _get(self, endpoint: str, **params):
        r = requests.get(f"{YADISK_API}/{endpoint}", headers=self.headers, params=params)
        r.raise_for_status()
        return r.json()

    def download(self, path: str) -> bytes:
        data = self._get("resources/download", path=path)
        r = requests.get(data["href"])
        r.raise_for_status()
        return r.content

    def upload(self, path: str, content: bytes, overwrite: bool = True):
        data = self._get("resources/upload", path=path, overwrite=str(overwrite).lower())
        r = requests.put(data["href"], data=content)
        r.raise_for_status()

    def mkdir(self, path: str):
        r = requests.put(
            f"{YADISK_API}/resources",
            headers=self.headers,
            params={"path": path},
        )
        if r.status_code not in (201, 409):
            r.raise_for_status()


def _default_folder() -> str:
    now = datetime.date.today()
    return f"Контент {MONTHS_RU[now.month - 1]} {now.year}"


def get_content_plan(yadisk: YaDiskClient, plan_path: str):
    raw = yadisk.download(plan_path)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb.active

    header_row = [str(c.value or "").strip() for c in ws[1]]
    col_map = {}
    for idx, h in enumerate(header_row):
        for field in _COL_ALIASES:
            if field not in col_map and _match_col(h, field):
                col_map[field] = idx

    tasks = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        vals = [str(c.value or "").strip() for c in row]
        if not any(vals):
            continue

        status = vals[col_map["status"]].lower() if "status" in col_map else ""
        if status in ("✓", "done", "выполнено", "+", "да", "yes"):
            continue

        topic = vals[col_map["topic"]] if "topic" in col_map else ""
        if not topic:
            continue

        tasks.append({
            "row_index": row_idx,
            "topic":       topic,
            "description": vals[col_map["description"]] if "description" in col_map else "",
            "contacts":    vals[col_map["contacts"]]    if "contacts"    in col_map else "",
            "key_aspects": vals[col_map["key_aspects"]] if "key_aspects" in col_map else "",
            "folder":      vals[col_map["folder"]]      if "folder"      in col_map and vals[col_map["folder"]] else _default_folder(),
        })

    return tasks, wb, col_map


def generate_funnel(client: Anthropic, context: str, skill: str, task: dict) -> str:
    prompt = f"""{skill}

---

## КОНТЕКСТ ПРОЕКТА

{context}

---

## ЗАДАНИЕ ИЗ КОНТЕНТ-ПЛАНА

**Тема воронки:** {task['topic']}
**Описание / бриф:** {task['description']}
**Контактные данные для CTA:** {task['contacts']}
**Ключевые аспекты и фокус:** {task['key_aspects']}

---

Сгенерируй полный пакет контентной воронки по всем частям скилла (1–9).
Используй данные о проекте из раздела «КОНТЕКСТ ПРОЕКТА».
Оформи результат как структурированный Markdown-документ с чёткими разделами.
Каждый раздел начинай с заголовка второго уровня (##).
"""

    message = client.messages.create(
        model="claude-opus-4-7",
        max_tokens=16000,
        messages=[{"role": "user", "content": prompt}],
    )
    return message.content[0].text


def save_funnel(yadisk: YaDiskClient, folder_name: str, topic: str, content: str) -> str:
    safe_topic = topic[:60].replace("/", "-").replace("\\", "-").strip()
    base = f"disk:/{folder_name}"
    topic_dir = f"{base}/{safe_topic}"
    file_path = f"{topic_dir}/воронка.md"

    yadisk.mkdir(base)
    yadisk.mkdir(topic_dir)
    yadisk.upload(file_path, content.encode("utf-8"))
    return file_path


def mark_done(wb, row_index: int, col_map: dict, result_path: str):
    ws = wb.active
    if "status" in col_map:
        ws.cell(row=row_index, column=col_map["status"] + 1).value = "✓"
    if "result" in col_map:
        ws.cell(row=row_index, column=col_map["result"] + 1).value = result_path


def main():
    yadisk_token = os.environ["YANDEX_DISK_TOKEN"]
    plan_path = os.environ.get("CONTENT_PLAN_PATH", "disk:/Контент-план/контент-план.xlsx")
    process_all = os.environ.get("PROCESS_ALL", "false").lower() == "true"

    yadisk = YaDiskClient(yadisk_token)
    claude = Anthropic()

    context = load_context()
    print(f"Контекст загружен ({len(context)} символов)")

    skill = load_skill()

    tasks, wb, col_map = get_content_plan(yadisk, plan_path)
    if not tasks:
        print("Нет незавершённых задач в контент-плане.")
        return

    print(f"Задач к обработке: {len(tasks)}")
    to_process = tasks if process_all else tasks[:1]

    for task in to_process:
        print(f"\n→ «{task['topic']}»")
        funnel = generate_funnel(claude, context, skill, task)
        print(f"  Сгенерировано: {len(funnel)} символов")

        path = save_funnel(yadisk, task["folder"], task["topic"], funnel)
        print(f"  Сохранено: {path}")

        mark_done(wb, task["row_index"], col_map, path)

    buf = io.BytesIO()
    wb.save(buf)
    yadisk.upload(plan_path, buf.getvalue(), overwrite=True)
    print("\nКонтент-план обновлён ✓")


if __name__ == "__main__":
    main()
